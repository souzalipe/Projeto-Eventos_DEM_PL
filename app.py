import re
import sys
import time
import numpy as np
import pandas as pd
import customtkinter as ctk
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from decimal import Decimal, ROUND_HALF_UP
from tkinter import filedialog, messagebox
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Tuple, Optional, Union
 
# ---------------- CONFIGURAÇÕES ----------------

CARTEIRA_CSV = None
MOVIMENTO_COTISTAS_PATH = None
BALANCETE_XLSX = None
DEM_PL_IN = None
DEM_PL_OUT = None
 
 
BALANCETE_SHEET: Optional[Union[str, int]] = None
COL_CONTA = "V"
COL_SALDO = "K"
SAFE_SAVE_WITH_SUFFIX = True
# Somatórios por bloco (item 7)
CEL_BLOCO_ACOES      = "J34"
CEL_BLOCO_RENDA_FIXA = "J40"
CEL_BLOCO_RECEITAS   = "J45"
CEL_BLOCO_DESPESAS   = "J55"
TOTAL_CELLS = {CEL_BLOCO_ACOES, CEL_BLOCO_RENDA_FIXA, CEL_BLOCO_RECEITAS, CEL_BLOCO_DESPESAS}
BLOCOS_RECONHECIDOS = {
    "Ações e Opções": "ACOES",
    "Renda fixa e outros valores mobiliários": "RENDA_FIXA",
    "Demais receitas": "RECEITAS",
    "Demais despesas": "DESPESAS",
}

 
NUM_FMT_INT_MIL = "#,##0;(#,##0);-"
ALIGN_RIGHT = Alignment(horizontal="right")
# ---------------- FUNÇÕES ----------------
def excel_col_to_zero_based(col_letter: str) -> int:
    col_letter = col_letter.strip().upper()
    num = 0
    for ch in col_letter:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Coluna inválida: {col_letter}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num - 1
def _read_balancete_df(balancete_path: Path, sheet: Optional[Union[str, int]]) -> pd.DataFrame:
    result = pd.read_excel(balancete_path, sheet_name=0 if sheet is None else sheet)
    if isinstance(result, dict):
        result = next(iter(result.values()))
    return result
def build_account_map(balancete_path: Path, sheet, col_conta, col_saldo) -> Dict[str, float]:
    df = _read_balancete_df(balancete_path, sheet)
    idx_conta = excel_col_to_zero_based(col_conta)
    idx_saldo = excel_col_to_zero_based(col_saldo)
    s_conta = df.iloc[:, idx_conta].astype(str)
    s_saldo = pd.to_numeric(df.iloc[:, idx_saldo], errors="coerce").fillna(0.0)
    contas = s_conta.str.extract(r"(\d+)", expand=False)
    tmp = pd.DataFrame({"conta": contas, "saldo": s_saldo}).dropna(subset=["conta"])
    return tmp.groupby("conta")["saldo"].sum().to_dict()


def format_valor_milhares(valor: int) -> str:
    """Formata valor conforme regras: negativo entre parênteses, zero como '-', separador milhar com ponto."""
    if valor == 0:
        return "-"
    elif valor < 0:
        return f"({abs(valor):,})".replace(",", ".")
    else:
        return f"{valor:,}".replace(",", ".")


def normalize_text_for_accounts(s: str) -> str:
    s = s.replace("\xa0", " ")
    s = s.replace("R$", "").replace(".", "")
    s = s.replace("—", "+").replace("–", "+").replace("-", "+")
    s = s.replace(";", "+").replace(",", "+")
    s = s.replace("’", "'").replace("`", "'")
    s = re.sub(r"\s+", "", s)
    return s
def parse_accounts_from_cell(val) -> List[str]:
    if val is None:
        return []
    s = normalize_text_for_accounts(str(val)).replace('"', "").replace("'", "")
    return re.findall(r"\d+", s)
def should_replace_cell(val) -> bool:
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    return bool(re.search(r"\d", normalize_text_for_accounts(str(val))))
def safe_save_workbook(wb, path_out: Path):
    try:
        wb.save(path_out)
        return path_out
    except PermissionError:
        ts = time.strftime("%Y%m%d_%H%M%S")
        alt = path_out.with_name(f"{path_out.stem}_{ts}{path_out.suffix}")
        wb.save(alt)
        return alt
# ✅ Item 6 – arredonda por mil
def round_thousands_cell(value_reais: float) -> int:
    return int((Decimal(value_reais) / Decimal(1000)).quantize(Decimal("0"), rounding=ROUND_HALF_UP))
def apply_int_mil_format(cell):
    cell.number_format = "General"
    cell.number_format = NUM_FMT_INT_MIL
    cell.alignment = ALIGN_RIGHT
 
 # --- [NOVO] Utilitários para CNPJ --------------------------------------------
def only_digits(s: str) -> str:
    import re
    return re.sub(r"\D+", "", s or "")
 
def mask_cnpj(cnpj: str) -> str:
    """
    Aplica máscara 00.000.000/0000-00 se houver 14 dígitos;
    caso contrário, retorna o próprio valor sem máscara.
    """
    d = only_digits(cnpj)
    if len(d) != 14:
        return cnpj  # devolve como vier (evita falha se a origem estiver irregular)
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"
 
 
 
def extract_cnpj_digits(val) -> Optional[str]:
    """
    Extrai exatamente 14 dígitos de um valor vindo do Balancete, lidando com:
    - floats terminando com .0 (ex.: 43096339000146.0)
    - inteiros
    - strings (com ou sem máscara)
    Retorna a string de 14 dígitos ou None.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
 
    # Caso 1: inteiro "puro"
    if isinstance(val, (int, np.integer)):
        d = f"{val:d}"
        return d.zfill(14) if len(d) <= 14 else (d[-14:] if len(d) > 14 else d)
 
    # Caso 2: float
    if isinstance(val, float):
        # Se for .0, convertemos para int sem casas
        if float(val).is_integer():
            d = f"{int(val):d}"
            return d.zfill(14) if len(d) <= 14 else (d[-14:] if len(d) > 14 else d)
        # Se não for inteiro, extraímos só dígitos
        s = str(val)
        digits = only_digits(s)
        # tenta achar um bloco de 14 dígitos
        m = re.search(r"(\d{14})", digits)
        return m.group(1) if m else (digits if len(digits) == 14 else None)
 
    # Caso 3: str (ou outros)
    s = str(val).strip()
    digits = only_digits(s)
    # Tenta capturar exatamente 14 dígitos
    m = re.search(r"(\d{14})", digits)
    if m:
        return m.group(1)
    # fallback: se tiver mais de 14, ficar com os 14 primeiros
    if len(digits) >= 14:
        return digits[:14]
    return None
 
 
def mask_cnpj_from_value(val) -> Optional[str]:
    """Usa extract_cnpj_digits e aplica a máscara 00.000.000/0000-00."""
    digits = extract_cnpj_digits(val)
    if not digits or len(digits) != 14:
        return None
    return f"{digits[0:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:14]}"
 
 
def get_cnpj_from_balancete(balancete_path: Path, sheet: Optional[Union[str, int]] = None) -> Optional[str]:
    """
    Procura a coluna 'Cnpj' (case-insensitive). Se não houver, usa a coluna G (índice 6).
    Retorna já no formato: 'CNPJ: 00.000.000/0000-00'.
    """
    df = _read_balancete_df(balancete_path, sheet)
 
    # 1) tenta encontrar a coluna 'Cnpj'
    cnpj_col = None
    for col in df.columns:
        if str(col).strip().lower() == "cnpj":
            cnpj_col = col
            break
 
    # 2) escolhe a série (ou fallback para coluna G)
    series = None
    if cnpj_col is not None:
        series = df[cnpj_col]
    else:
        try:
            series = df.iloc[:, 6]  # G = índice 6 (A=0)
        except Exception:
            return None
 
    # 3) percorre até achar um valor válido
    for val in series:
        masked = mask_cnpj_from_value(val)
        if masked:
            return f"CNPJ: {masked}"
    return None


def formatar_ptbr(num: Decimal, casas: int = 3) -> str:
    """
    Formata número para PT-BR com separador de milhar '.' e decimal ','.
    Arredonda HALF_UP para 'casas' decimais e sempre exibe essas casas.
    """
    # Quantiza com arredondamento "meio para cima"
    q = Decimal(1).scaleb(-casas)             # ex.: 0.001 para 3 casas
    n = Decimal(num).quantize(q, rounding=ROUND_HALF_UP)

    # Usa formatação "en-US" e troca separadores
    s = f"{n:,.{casas}f}"                     # ex.: 123,519,889.535
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


# -----------------------------------------------------------------------------
 


# def get_last_ncotas(carteira_csv: Path) -> Optional[str]:
#     """
#     Lê 'Carteira Diária', usa o header da penúltima linha como cabeçalho,
#     pega o último valor válido da coluna 'NCotas', e devolve como texto
#     com pontuação de milhar (.) e 3 casas decimais (,) — ex.: 123.519.889,535.
#     """
#     import io

#     # 1) Ler o arquivo "cru" com encoding correto e sem assumir header
#     with open(carteira_csv, "r", encoding="latin-1", errors="ignore") as f:
#         raw_lines = [ln.rstrip("\n") for ln in f if ln.strip()]

#     if len(raw_lines) < 2:
#         return None

#     # 2) A penúltima linha é o header; a última linha geralmente é rodapé/outro bloco
#     header_line = raw_lines[-2]
#     data_lines = raw_lines[:-2]  # todas as linhas antes do header

#     # 3) Monta um CSV em memória com o header na primeira linha + os dados abaixo
#     csv_text = "\n".join([header_line] + data_lines)

#     # 4) Lê com separador ';' e decimal ',' (isso já converte strings numéricas corretamente)
#     df = pd.read_csv(io.StringIO(csv_text),
#                      sep=";",
#                      decimal=",",
#                      engine="python",
#                      encoding="latin-1")

#     # 5) Garante que a coluna existe (atenção a espaços/maiúsculas/minúsculas)
#     #    Normaliza nomes de colunas tirando espaços extras
#     df.columns = [str(c).strip() for c in df.columns]
#     if "NCotas" not in df.columns:
#         # Tenta variações comuns
#         for c in df.columns:
#             if c.lower().strip() == "ncotas":
#                 df.rename(columns={c: "NCotas"}, inplace=True)
#                 break
#         if "NCotas" not in df.columns:
#             print("⚠️ Coluna 'NCotas' não encontrada no Carteira Diária.")
#             return None

#     # 6) Converte a coluna para número (pandas já entende decimal=',', mas reforçamos)
#     serie = pd.to_numeric(df["NCotas"], errors="coerce")
#     serie = serie.dropna()
#     if serie.empty:
#         return None

#     # 7) Último valor válido (última linha com número)
#     num = float(serie.iloc[-1])

#     # 8) Formata com 3 casas decimais e milhar com ponto
#     #    Primeiro cria em notação US (123,519,889.535), depois troca separadores:
#     texto = f"{num:,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
#     return texto



def get_last_ncotas(carteira_csv: Path) -> Optional[str]:
    """
    Lê o arquivo Carteira Diária, ajusta o header (penúltima linha),
    e retorna o último valor da coluna 'NCotas' já formatado como texto
    PT-BR com 3 casas decimais (ex.: '123.519.889,535').
    """
    import io
    # 1) Ler linhas cruas com encoding correto
    with open(carteira_csv, "r", encoding="latin-1", errors="ignore") as f:
        raw_lines = [ln.rstrip("\n") for ln in f if ln.strip()]
    if len(raw_lines) < 2:
        return None

    # 2) Header na penúltima linha; dados = todas as anteriores
    header_line = raw_lines[-2]
    data_lines  = raw_lines[:-2]

    # 3) Montar CSV em memória com header na primeira linha
    csv_text = "\n".join([header_line] + data_lines)

    # 4) Ler em pandas respeitando separador ';' e decimal ','
    df = pd.read_csv(
        io.StringIO(csv_text),
        sep=";",
        decimal=",",
        engine="python",
        encoding="latin-1",
    )

    # 5) Normalizar nome da coluna e garantir 'NCotas'
    df.columns = [str(c).strip() for c in df.columns]
    if "NCotas" not in df.columns:
        for c in df.columns:
            if c.lower().strip() == "ncotas":
                df.rename(columns={c: "NCotas"}, inplace=True)
                break
    if "NCotas" not in df.columns:
        print("⚠️ Coluna 'NCotas' não encontrada no Carteira Diária.")
        return None

    # 6) Converter a série para número (pandas já entende decimal=',')
    serie = pd.to_numeric(df["NCotas"], errors="coerce").dropna()
    if serie.empty:
        return None

    # 7) Último valor válido
    num = Decimal(str(serie.iloc[-1]))

    # 8) Formatar com 3 casas no padrão PT-BR
    return formatar_ptbr(num, casas=3)




def get_last_vlcotas(carteira_csv: Path) -> Optional[str]:
    """
    Lê o arquivo Carteira Diária, usa o header da penúltima linha como cabeçalho,
    pega o último valor válido da coluna 'VlCotas' e retorna em texto com 6 casas decimais.
    Ex.: 123519889.535287 -> "123519889.535287" (ou com vírgula, se quiser trocar depois).
    """
    import io

    # 1) Ler linhas cruas com encoding latino (evita UnicodeDecodeError em Ç/Á/etc.)
    with open(carteira_csv, "r", encoding="latin-1", errors="ignore") as f:
        raw_lines = [ln.rstrip("\n") for ln in f if ln.strip()]

    if len(raw_lines) < 2:
        return None

    # 2) Header na penúltima linha; dados = todas as anteriores
    header_line = raw_lines[-2]
    data_lines = raw_lines[:-2]

    # 3) Montar CSV com header na primeira linha
    csv_text = "\n".join([header_line] + data_lines)

    # 4) Ler em pandas respeitando separador e decimal
    #    Aqui usamos decimal="," porque os números do arquivo vêm no padrão PT-BR
    df = pd.read_csv(
        io.StringIO(csv_text),
        sep=";",
        decimal=",",
        engine="python",
        encoding="latin-1",
    )

    # 5) Normalizar nomes de colunas (tira espaços)
    df.columns = [str(c).strip() for c in df.columns]

    # 6) Garantir a coluna 'VlCotas'
    if "VlCotas" not in df.columns:
        # Tenta variações de letra maiuscula/minuscula/espacos
        for c in df.columns:
            if c.lower().strip() == "vlcotas":
                df.rename(columns={c: "VlCotas"}, inplace=True)
                break
        if "VlCotas" not in df.columns:
            print("⚠️ Coluna 'VlCotas' não encontrada no Carteira Diária.")
            return None

    # 7) Converter para número (pandas já entende decimal=',')
    serie = pd.to_numeric(df["VlCotas"], errors="coerce").dropna()
    if serie.empty:
        return None

    # 8) Pegar o último valor válido
    num = float(serie.iloc[-1])

    # 9) Formatar com 6 casas decimais
    #    Se quiser vírgula como decimal, troque por .replace(".", ",")
    return f"{num:.6f}"


 
from pathlib import Path
from typing import Dict, Optional
from openpyxl import load_workbook
 
def replace_in_dem_pl(dem_in: Path,dem_out: Path,acc_map: Dict[str, float],cnpj_str: Optional[str] = None) -> Path:
    """
    Abre o modelo Dem-PL, percorre todas as abas substituindo células que contenham
    referências a contas do balancete pelas somas correspondentes (em milhares, com
    arredondamento half-up), acumula os totais por bloco e atualiza:
      - J34 (Ações e Opções)
      - J40 (Renda fixa e outros valores mobiliários)
      - J45 (Demais receitas)
      - J55 (Demais despesas)
      - J58 (Total geral = soma de J34+J40+J45+J55)
      - L8  (CNPJ prefixado: "CNPJ: 00.000.000/0000-00", como texto)
 
    Ao final, salva com 'safe_save_workbook' (que lida com arquivo bloqueado)
    e retorna o caminho efetivo gerado.
 
    Parâmetros
    ----------
    dem_in : Path
        Caminho do arquivo modelo de entrada (.xlsx).
    dem_out : Path
        Caminho desejado para o arquivo de saída (.xlsx).
    acc_map : Dict[str, float]
        Mapa {codigo_conta: saldo_em_reais} extraído do balancete.
    cnpj_str : Optional[str]
        Texto já formatado do CNPJ (ex.: "CNPJ: 00.000.000/0000-00"). Se None, não escreve.
    """
    # Abre o workbook do modelo
    wb = load_workbook(dem_in, data_only=False)
 
    # Acumuladores opcionais (mantidos caso você queira usar no futuro)
    changes = []
    totals_por_conta = {}
    missing_codes = {}
 
    # Somatórios por bloco (em milhares)
    soma_blocos = {"ACOES": 0, "RENDA_FIXA": 0, "RECEITAS": 0, "DESPESAS": 0}
    bloco_atual = None
 
    # ---------------------------
    # Passo 1: Varre TODAS as abas e substitui as células "calculáveis"
    # ---------------------------
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            # Verifica o bloco pela coluna A (primeira coluna)
            col_a_val = row[0].value
            if isinstance(col_a_val, str):
                key = col_a_val.strip()
                if key in BLOCOS_RECONHECIDOS:
                    bloco_atual = BLOCOS_RECONHECIDOS[key]
 
            for cell in row:
                # Não mexer nas células de total consolidadas
                if cell.coordinate.upper() in TOTAL_CELLS:
                    continue
 
                # Só tentamos substituir se a célula "parece" conter contas
                if not should_replace_cell(cell.value):
                    continue
 
                raw_expr = str(cell.value)
                contas = parse_accounts_from_cell(raw_expr)
                if not contas:
                    continue
 
                # Soma em reais das contas existentes no mapa
                total_reais = 0.0
                for c in contas:
                    v = float(acc_map.get(c, 0.0))
                    total_reais += v
                    totals_por_conta[c] = totals_por_conta.get(c, 0.0) + v
                    if c not in acc_map:
                        missing_codes[c] = missing_codes.get(c, 0) + 1
 
                # Converte para inteiro em milhares, com arredondamento HALF_UP
                val_mil = round_thousands_cell(total_reais)
 
                # Escreve o valor e aplica formatação
                cell.value = val_mil
                apply_int_mil_format(cell)
 
                # Log opcional da mudança
                changes.append((f"{ws.title}!{cell.coordinate}", raw_expr, total_reais, val_mil))
 
                # Acumula no bloco atual (se estivermos dentro de um bloco reconhecido)
                if bloco_atual in soma_blocos:
                    soma_blocos[bloco_atual] += val_mil
 
    # ---------------------------
    # Passo 2: Atualiza a 1ª ABA com os somatórios e o total geral
    # ---------------------------
    ws0 = wb.worksheets[0]
 
    # Preenche cada célula de bloco
    for coord, key in [
        (CEL_BLOCO_ACOES, "ACOES"),
        (CEL_BLOCO_RENDA_FIXA, "RENDA_FIXA"),
        (CEL_BLOCO_RECEITAS, "RECEITAS"),
        (CEL_BLOCO_DESPESAS, "DESPESAS"),
    ]:
        ws0[coord].value = soma_blocos[key]
        apply_int_mil_format(ws0[coord])
 
    # Total geral (J58) = soma dos inteiros em milhares
    CEL_TOTAL_GERAL = "J58"
    total_geral = (
        soma_blocos["ACOES"]
        + soma_blocos["RENDA_FIXA"]
        + soma_blocos["RECEITAS"]
        + soma_blocos["DESPESAS"]
    )
    ws0[CEL_TOTAL_GERAL].value = total_geral
    apply_int_mil_format(ws0[CEL_TOTAL_GERAL])


    # --- NOVO: incluir conta 61180 na célula J23 ---
    CONTA_EXTRA = "61180"
    CEL_EXTRA = "J23"

    # Busca saldo da conta no mapa acc_map (já carregado do balancete)
    saldo_reais = float(acc_map.get(CONTA_EXTRA, 0.0))

    # Converte para milhares e arredonda
    valor_mil = round_thousands_cell(saldo_reais)

    # Formata conforme regra
    valor_formatado = format_valor_milhares(valor_mil)

    # Escreve na célula J23
    ws0[CEL_EXTRA].value = valor_formatado
    ws0[CEL_EXTRA].number_format = "@"


    
    # --- NOVO: preencher D18 com NCotas da Carteira Diária ---
    
    if CARTEIRA_CSV.exists():
        val_d18 = get_last_ncotas(CARTEIRA_CSV)
        if val_d18:
            ws0["D18"].value = val_d18
            print(f"[DEBUG] Valor NCotas formatado para D18: {val_d18}")
            ws0["D18"].number_format = "@"

 
    # ---------------------------
    # Passo 3: CNPJ em L8 (como texto), se informado
    # ---------------------------
    if cnpj_str:
        ws0["L8"].value = str(cnpj_str)      # garante string
        ws0["L8"].number_format = "@"        # força formato TEXTO
        try:
            ws0["L8"].alignment = ALIGN_RIGHT
        except Exception:
            pass
 
    # ---------------------------
    # Passo 4: Salvar e retornar o caminho efetivo
    # ---------------------------
    path_saida = safe_save_workbook(wb, dem_out)
 

def preencher_movimento_cotistas(dem_out: Path, mov_path: Path):
    """
    Lê o arquivo Movimento de Cotistas (CSV), ajusta cabeçalho, extrai os últimos valores
    das colunas NCATOT_Tot e NCRTOT_Tot, formata e escreve nas células D20 e D22 do Excel.
    """
    try:
        # 1. Verificar se os arquivos existem
        if not mov_path.exists():
            print(f"ERRO — Arquivo Movimento de Cotistas não encontrado: {mov_path}")
            return
        if not dem_out.exists():
            print(f"ERRO — Arquivo Dem_PL_Modelo_preenchido não encontrado: {dem_out}")
            return

        # 2. Ler todas as linhas do CSV
        with open(mov_path, 'r', encoding='latin1') as f:
            linhas = f.readlines()

        if len(linhas) < 2:
            print("ERRO — Arquivo Movimento de Cotistas está vazio ou inválido.")
            return

        # 3. Ajustar cabeçalho (última linha vira header)
        header = linhas[-1].strip().split(';')
        dados = linhas[:-1]

        # 4. Criar DataFrame com pandas
        df_mov = pd.DataFrame([linha.strip().split(';') for linha in dados], columns=header)

        # 5. Verificar colunas
        if 'NCATOT_Tot' not in df_mov.columns or 'NCRTOT_Tot' not in df_mov.columns:
            print("ERRO — Colunas NCATOT_Tot ou NCRTOT_Tot não encontradas no arquivo.")
            return

        # 6. Extrair últimos valores
        valor_ncatot = df_mov['NCATOT_Tot'].dropna().iloc[-1]
        valor_ncrtot = df_mov['NCRTOT_Tot'].dropna().iloc[-1]

        # 7. Converter para float (tratando vírgula e ponto)
        valor_ncatot = float(str(valor_ncatot).replace('.', '').replace(',', '.'))
        valor_ncrtot = float(str(valor_ncrtot).replace('.', '').replace(',', '.'))

        # 8. Formatar padrão brasileiro (milhar com ponto, decimal com vírgula)
        def formatar(valor):
            return f"{valor:,.3f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        valor_formatado_ncatot = formatar(valor_ncatot)
        valor_formatado_ncrtot = formatar(valor_ncrtot)

        # 9. Abrir Excel e escrever nas células D20 e D22
        wb = load_workbook(dem_out)
        ws = wb.worksheets[0]

        ws['D20'].value = valor_formatado_ncatot
        ws['D22'].value = valor_formatado_ncrtot

        ws['D20'].number_format = '@'
        ws['D22'].number_format = '@'

        # 10. Salvar arquivo
        wb.save(dem_out)

        print("[OK] Valores inseridos com sucesso:")
        print(f"D20 (NCATOT_Tot): {valor_formatado_ncatot}")
        print(f"D22 (NCRTOT_Tot): {valor_formatado_ncrtot}")

    except Exception as e:
        print(f"ERRO — Falha ao preencher Movimento de Cotistas: {e}")

# ----- Interface -------------

def abrir_interface():
    import customtkinter as ctk
    from tkinter import filedialog, messagebox
    from pathlib import Path

    # Configuração do tema
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    # Função para selecionar arquivo
    def selecionar_arquivo(entry_widget):
        caminho = filedialog.askopenfilename(title="Selecione o arquivo", filetypes=[("Todos os arquivos", "*.*")])
        if caminho:
            entry_widget.delete(0, "end")
            entry_widget.insert(0, caminho)

    # Função para executar processamento
    def executar():
        balancete = entry_balancete.get().strip()
        dem_pl_in = entry_dem_pl.get().strip()
        movimento = entry_movimento.get().strip()
        carteira = entry_carteira.get().strip()

        # Validação dos arquivos obrigatórios
        if not all([balancete, dem_pl_in, movimento]):
            messagebox.showerror("Erro", "Por favor, selecione todos os arquivos obrigatórios.")
            return

        # Atualiza variáveis globais como Path
        global CARTEIRA_CSV, MOVIMENTO_COTISTAS_PATH, BALANCETE_XLSX, DEM_PL_IN, DEM_PL_OUT
        
        CARTEIRA_CSV = Path(carteira) if carteira else None
        MOVIMENTO_COTISTAS_PATH = Path(movimento)
        BALANCETE_XLSX = Path(balancete)
        DEM_PL_IN = Path(dem_pl_in)
        DEM_PL_OUT = Path("Dem_PL_Modelo_preenchido.xlsx")  # Pode manter fixo ou permitir escolha

        # Executa main()
        try:
            main()
            messagebox.showinfo("Sucesso", f"Processamento concluído!\nArquivo gerado: {DEM_PL_OUT}")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

    # Criar janela principal
    janela = ctk.CTk()
    janela.title("Processador Dem-PL")
    janela.geometry("759x390")

    # Labels e campos
    label_balancete = ctk.CTkLabel(janela, text="Balancete XLSX:")
    label_balancete.grid(row=0, column=0, padx=10, pady=10, sticky="w")
    entry_balancete = ctk.CTkEntry(janela, width=400)
    entry_balancete.grid(row=0, column=1, padx=10, pady=10)
    btn_balancete = ctk.CTkButton(janela, text="Selecionar", command=lambda: selecionar_arquivo(entry_balancete))
    btn_balancete.grid(row=0, column=2, padx=10, pady=10)

    label_dem_pl = ctk.CTkLabel(janela, text="Dem-PL Modelo XLSX:")
    label_dem_pl.grid(row=1, column=0, padx=10, pady=10, sticky="w")
    entry_dem_pl = ctk.CTkEntry(janela, width=400)
    entry_dem_pl.grid(row=1, column=1, padx=10, pady=10)
    btn_dem_pl = ctk.CTkButton(janela, text="Selecionar", command=lambda: selecionar_arquivo(entry_dem_pl))
    btn_dem_pl.grid(row=1, column=2, padx=10, pady=10)

    label_movimento = ctk.CTkLabel(janela, text="Movimento de Cotistas CSV:")
    label_movimento.grid(row=2, column=0, padx=10, pady=10, sticky="w")
    entry_movimento = ctk.CTkEntry(janela, width=400)
    entry_movimento.grid(row=2, column=1, padx=10, pady=10)
    btn_movimento = ctk.CTkButton(janela, text="Selecionar", command=lambda: selecionar_arquivo(entry_movimento))
    btn_movimento.grid(row=2, column=2, padx=10, pady=10)

    label_carteira = ctk.CTkLabel(janela, text="Carteira CSV (opcional):")
    label_carteira.grid(row=3, column=0, padx=10, pady=10, sticky="w")
    entry_carteira = ctk.CTkEntry(janela, width=400)
    entry_carteira.grid(row=3, column=1, padx=10, pady=10)
    btn_carteira = ctk.CTkButton(janela, text="Selecionar", command=lambda: selecionar_arquivo(entry_carteira))
    btn_carteira.grid(row=3, column=2, padx=10, pady=10)

    # Botão executar
    btn_executar = ctk.CTkButton(janela, text="Gerar Documento", command=executar, fg_color="green", text_color="white")
    btn_executar.grid(row=4, column=1, pady=20)

    janela.mainloop()




def main():
    bal = Path(BALANCETE_XLSX)
    dem_in = Path(DEM_PL_IN)
    mov_path = Path(MOVIMENTO_COTISTAS_PATH)


    if not bal.exists():
        print("ERRO — Balancete não encontrado:", bal)
        sys.exit(1)
    if not dem_in.exists():
        print("ERRO — Modelo não encontrado:", dem_in)
        sys.exit(1)

    # 1) mapa de contas
    acc_map = build_account_map(bal, BALANCETE_SHEET, COL_CONTA, COL_SALDO)

    # 2) captura CNPJ
    cnpj_str = get_cnpj_from_balancete(bal, BALANCETE_SHEET)

    # 3) executa preenchimento Dem-PL
    out_file = replace_in_dem_pl(dem_in, Path(DEM_PL_OUT), acc_map, cnpj_str)

    # 4) executa preenchimento Movimento de Cotistas (D20 e D22)
    preencher_movimento_cotistas(Path(DEM_PL_OUT), mov_path)

    print("\n[ OK ] Concluído!")
    print(f"Arquivo gerado: {out_file}")
    if cnpj_str:
        print(f"CNPJ escrito em L8: {cnpj_str}")
    else:
        print("Não foi possível localizar CNPJ no balancete (coluna 'Cnpj' ou G).")



if __name__ == "__main__":
    abrir_interface()
